"""
Generate the AI Audit Log Excel workbook for Lab 4 (DBI202),
following AI_AuditLog_Student_Guideline_DB.txt and based on Lab4_report.docx.

Sheets:
  1. Metadata & Summary
  2. Detailed AI Audit Log
  3. Hallucination Detection Log

Run:  python generate_ai_auditlog.py
Output: Lab4_AI_AuditLog.xlsx
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---- palette
TITLE_BLUE = "1F4E79"
HEADER_BLUE = "2E75B6"
SECTION_BLUE = "1F4E79"
YELLOW = "FFFF00"
HDR_FILL = PatternFill("solid", fgColor=HEADER_BLUE)
YEL_FILL = PatternFill("solid", fgColor=YELLOW)
WHITE = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
SECTION_FONT = Font(color=SECTION_BLUE, bold=True, size=12, name="Calibri")
TITLE_FONT = Font(color=TITLE_BLUE, bold=True, size=16, name="Calibri")
BOLD = Font(bold=True, size=11, name="Calibri")
NORMAL = Font(size=11, name="Calibri")
ITALIC = Font(size=10, italic=True, color="808080", name="Calibri")
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
TOPLEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFTV = Alignment(horizontal="left", vertical="center", wrap_text=True)

STUDENT = {
    "Student Name:": "Nguyễn Phú Khương",
    "Student ID:": "SE203056",
    "Course:": "DBI202",
    "Assignment:": "Lab4 - Relational Database Design Process",
}

TOTAL_PROMPTS = 40
CORE_PROMPTS = 8
HALLUCINATIONS = 4

AI_TOOLS = [
    ["ChatGPT", "Decision-making on data model & normalization (relational vs NoSQL, "
                "supertype/subtype, BCNF)", "High", "Critical thinking support"],
    ["GitHub Copilot", "Cross-checking constraints, FK referential actions and SQL Server syntax",
     "Medium", "Verification support"],
    ["Gemini", "Report structure, wording and diagram design", "Medium", "Support with report design"],
]

DTC = [
    ["Decomposition", 2, "\u2265 1"],
    ["Pattern Recognition", 2, "\u2265 1"],
    ["Abstraction", 2, "\u2265 1"],
    ["Algorithms", 2, "\u2265 1"],
]


def hd(ct, cx, cs, do):
    return (f"Critical Thinking: {ct}\n"
            f"Contextualization: {cx}\n"
            f"Creative Synthesis: {cs}\n"
            f"Decision Ownership: {do}")


# Detailed log entries: [Entry#, Type, Stage, Problem, Prompt, AI Response, Human Delta, Evidence]
ENTRIES = [
    ["001", "DECISION-MAKING", "Decomposition",
     "Need to choose the overall data architecture for the Household Cleaning Robot Sales & "
     "Maintenance system (relational vs NoSQL vs hybrid).",
     "\"For a system handling robot sales, customer records, warranty, maintenance history and "
     "IoT device error logs, should I use a relational model, NoSQL, or a hybrid?\"",
     "AI proposed a hybrid: relational DB for sales/warranty plus a NoSQL document store for IoT "
     "error logs, citing flexibility and horizontal scalability for high-volume logs.",
     hd("The hybrid adds operational complexity and breaks the single-schema requirement; our data "
        "is highly structured and transactional, so a split store is not justified.",
        "This is a DBI202 lab that requires ONE normalized relational schema; maintenance reports "
        "need joins between device logs and robot/customer records.",
        "I folded IoT data into relational tables (DeviceLog + LogStatistic) keeping FK integrity to "
        "RobotUnit, instead of an external document store.",
        "Chose a purely relational model of 16 relations so device logs join directly with robot and "
        "customer data."),
     "Lab4_report Section 2 & 3 (16 entities, relational schema)"],

    ["002", "DECISION-MAKING", "Decomposition",
     "A payment can be for a sales order OR a maintenance service. How to model it without "
     "ambiguous nullable foreign keys?",
     "\"How should I model a Payment that is either for a sales order or for a maintenance service "
     "in a relational schema?\"",
     "AI suggested a single Payment table containing two nullable foreign keys: OrderID and "
     "ServiceRecordID, filling whichever applies.",
     hd("Two nullable FKs allow invalid rows (a payment linked to both, or to neither), which "
        "violates the rule that each payment maps to exactly one transaction.",
        "Business rule: every sales transaction and every service operation has its own payment "
        "record; the two cases are mutually exclusive.",
        "Applied a supertype/subtype pattern: a common Payment table plus OrderPayment and "
        "ServicePayment subtype tables, each with a mandatory (NOT NULL) FK.",
        "Chose supertype/subtype decomposition to eliminate nullable FKs and enforce exclusivity."),
     "Lab4_report Section 3.3; relations Payment / OrderPayment / ServicePayment"],

    ["003", "PROBLEM-SOLVING", "Pattern Recognition",
     "RobotModel has many features, MaintenanceRecord has many replaced parts, DeviceLog has many "
     "usage metrics. How to store these repeating attributes?",
     "\"How to store multi-valued attributes such as a list of features for a product in a "
     "relational database?\"",
     "AI suggested storing the list as a single comma-separated string column, or as a JSON column, "
     "to keep the table simple.",
     hd("CSV/JSON columns violate 1NF, prevent filtering/searching by an individual value and break "
        "the normalization required by the lab.",
        "Customers must browse robots BY feature and reports must analyse parts per record, so the "
        "values must be individually queryable.",
        "Recognized the same repeating-group pattern in 3 places and decomposed each into a weak "
        "entity with a composite key (ModelFeature, ReplacedPart, LogStatistic).",
        "Chose three separate normalized tables (1NF/BCNF) over a CSV/JSON column."),
     "Lab4_report Section 3.3; ModelFeature / ReplacedPart / LogStatistic"],

    ["004", "DECISION-MAKING", "Pattern Recognition",
     "Each robot unit has exactly one warranty registration (1:1). How to enforce this cardinality?",
     "\"How do I enforce a 1:1 relationship between RobotUnit and WarrantyRegistration in SQL "
     "Server?\"",
     "AI recommended placing a WarrantyID foreign key directly inside the RobotUnit table to make "
     "the link one-to-one.",
     hd("Embedding the warranty FK in RobotUnit forces every unit to depend on a warranty row and "
        "complicates inventory inserts before a unit is sold/registered.",
        "A unit can sit in inventory unsold, so warranty registration must be optional until the "
        "robot is actually registered by a customer.",
        "Kept WarrantyRegistration as its own table and enforced 1:1 with a UNIQUE constraint on "
        "RobotID, allowing optional registration.",
        "Chose UNIQUE(RobotID) on WarrantyRegistration instead of an FK inside RobotUnit."),
     "Lab4_report Section 5 UNIQUE constraints (WarrantyRegistration.RobotID)"],

    ["005", "DECISION-MAKING", "Abstraction",
     "Decide the target normal form for the 16 relations.",
     "\"Is 3NF enough for this schema or should I normalize to BCNF? What are the trade-offs?\"",
     "AI stated that 3NF is generally enough and that going to BCNF can hurt performance, so I "
     "should stop at 3NF.",
     hd("The blanket 'BCNF hurts performance' is an oversimplification; for our composite-key "
        "relations 3NF and BCNF differ, and the lab explicitly requires BCNF.",
        "Lab3 already established BCNF for this domain, so Lab4 must stay consistent with it.",
        "Re-derived the functional dependencies and confirmed every determinant is a candidate "
        "key, so BCNF holds with no lossy or dependency-breaking decomposition.",
        "Chose BCNF for all 16 relations and documented the reasoning."),
     "Lab4_report Section 3 (BCNF note) & Section 6 reflection"],

    ["006", "VERIFICATION", "Abstraction",
     "Listing the core integrity constraints/properties to describe in the report.",
     "\"What are the core constraints/properties I should mention for a relational data model?\"",
     "AI listed Domain, Key and Referential Integrity, but also added \"Inheritance constraints\" "
     "and \"Polymorphism\" as core relational properties.",
     hd("Inheritance and polymorphism are Object-Oriented Programming concepts, NOT native "
        "relational-model constraints - this is a fabrication / logic error.",
        "DBI202 relational theory defines only domain, key, entity and referential integrity "
        "constraints.",
        "Cross-checked against the DBI202 textbook and lecture notes and removed the fabricated "
        "items before writing Section 2.1 / 5.",
        "Kept only the valid relational integrity constraints in the report."),
     "Hallucination Log H1; Lab4_report Section 5 constraints"],

    ["007", "DECISION-MAKING", "Algorithms",
     "Choose ON DELETE / ON UPDATE referential actions for all foreign keys.",
     "\"Should I set ON DELETE CASCADE and ON UPDATE CASCADE on all foreign keys for automatic "
     "cleanup?\"",
     "AI said yes - apply CASCADE on every FK so child rows are cleaned up and updated "
     "automatically.",
     hd("Cascading every FK triggers the SQL Server error 'may cause cycles or multiple cascade "
        "paths' (e.g. Customer -> SalesOrder and Customer -> OrderPayment via SalesOrder).",
        "SQL Server forbids more than one cascade path that reaches the same table.",
        "Designed a mixed policy: CASCADE only on true ownership/child tables (junction tables, "
        "order lines) and NO ACTION on shared references.",
        "Chose per-FK referential actions documented in the FK table to avoid cascade conflicts."),
     "Hallucination Log H2; Lab4_report Section 5.2 FK table"],

    ["008", "PROBLEM-SOLVING", "Algorithms",
     "Choose SQL Server data types and validation rules for money, text and status columns.",
     "\"Best SQL Server data types for prices, descriptions and status fields, and how should I "
     "validate them?\"",
     "AI suggested MONEY for prices, VARCHAR for all text fields, and said no special validation "
     "was needed.",
     hd("MONEY has rounding/precision issues, VARCHAR drops Unicode (Vietnamese names/addresses), "
        "and with no CHECK the DB would accept negative prices and invalid statuses.",
        "Vietnamese customer names and addresses require Unicode; prices must be exact and "
        "positive; statuses are a fixed enumeration.",
        "Used DECIMAL(18,2) for money, NVARCHAR for human text, VARCHAR only for codes, plus CHECK "
        "constraints (price > 0, status IN (...)) and IDENTITY for surrogate keys.",
        "Chose DECIMAL + NVARCHAR + CHECK constraints over the AI defaults."),
     "Hallucination Log H3 & H4; Lab4_report Section 4 data types & Section 5 CHECK constraints"],
]

# Hallucination log: [Entry#, Type, AI's Claim, Reality Check, How Detected, Corrective Action]
HALLU = [
    ["006", "Fabrication / Logic Error",
     "Relational Data Model includes \"Inheritance constraints\" and \"Polymorphism\" as core "
     "constraints/properties.",
     "The relational model does NOT support inheritance or polymorphism natively; these are "
     "Object-Oriented Programming concepts.",
     "Cross-checked with the DBI202 textbook, relational-theory lectures and academic materials on "
     "data modeling.",
     "Removed the fabricated concepts and kept only valid relational constraints (Domain, Key, "
     "Entity Integrity, Referential Integrity) in Section 2.1 / 5."],

    ["007", "Logic Error / Context Misunderstanding",
     "Setting ON DELETE CASCADE and ON UPDATE CASCADE on every foreign key is safe and recommended.",
     "SQL Server rejects schemas with multiple cascade paths to the same table; e.g. Customer "
     "cascading through both SalesOrder and OrderPayment raises error 1785.",
     "Reasoned through the FK graph and confirmed against SQL Server documentation that multiple "
     "cascade paths are not allowed.",
     "Replaced the blanket policy with a mixed CASCADE / NO ACTION design documented in the FK "
     "constraint table (Section 5.2)."],

    ["008", "Outdated Info",
     "Use the MONEY data type for all price/amount columns in SQL Server.",
     "MONEY suffers from rounding errors in division/aggregation and is discouraged for financial "
     "precision; DECIMAL(p,s) is the recommended modern choice.",
     "Compared MONEY vs DECIMAL behaviour and checked current SQL Server best-practice guidance.",
     "Used DECIMAL(18,2) for UnitPrice, TotalAmount, SellingPrice, Amount and ServiceFee "
     "(Section 4)."],

    ["003", "Oversimplification",
     "Multi-valued attributes (features, replaced parts, usage metrics) can be stored as a single "
     "comma-separated / JSON column.",
     "That violates 1NF and prevents querying/filtering by an individual value, breaking the "
     "normalization the lab requires.",
     "Tested the query requirement 'browse robots by feature' against a CSV column and saw it "
     "cannot be indexed/filtered cleanly.",
     "Decomposed each repeating group into its own normalized table (ModelFeature, ReplacedPart, "
     "LogStatistic) with composite primary keys."],
]


def style_title(ws, text, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=text)
    c.font = TITLE_FONT
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26


def header_row(ws, row, headers, widths=None):
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=j, value=h)
        c.fill = HDR_FILL
        c.font = WHITE
        c.alignment = CENTER
        c.border = BORDER
    if widths:
        for j, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(j)].width = w


def build_summary(wb):
    ws = wb.active
    ws.title = "Metadata & Summary"
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCD", [30, 42, 22, 30]):
        ws.column_dimensions[col].width = w
    style_title(ws, "AI AUDIT LOG - METADATA & SUMMARY", 4)

    r = 3
    ws.cell(row=r, column=1, value="STUDENT INFORMATION").font = SECTION_FONT
    r += 1
    for k, v in STUDENT.items():
        ws.cell(row=r, column=1, value=k).font = BOLD
        ws.cell(row=r, column=3, value=v).font = NORMAL
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="AI USAGE SUMMARY").font = SECTION_FONT
    r += 1
    summary = [
        ("Total Prompts Used (all AI tools):", TOTAL_PROMPTS, False, ""),
        ("Core Prompts Logged:", CORE_PROMPTS, True, "Range for DBI202: 5 - 8"),
        ("Selection Ratio:", f"{CORE_PROMPTS / TOTAL_PROMPTS * 100:.1f}%", False, "Should be 10-20%"),
        ("Hallucination Detected:", HALLUCINATIONS, True, "Required for DBI202: \u2265 3"),
    ]
    for label, val, hl, note in summary:
        ws.cell(row=r, column=1, value=label).font = BOLD
        c = ws.cell(row=r, column=3, value=val)
        c.font = BOLD
        c.alignment = Alignment(horizontal="center")
        if hl:
            c.fill = YEL_FILL
        if note:
            n = ws.cell(row=r, column=4, value=note)
            n.font = ITALIC
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="AI TOOLS USED").font = SECTION_FONT
    r += 1
    header_row(ws, r, ["AI Tool", "Purpose", "Frequency", "Main Value"])
    r += 1
    for tool in AI_TOOLS:
        for j, val in enumerate(tool, start=1):
            c = ws.cell(row=r, column=j, value=val)
            c.font = NORMAL
            c.alignment = TOPLEFT
            c.border = BORDER
        ws.row_dimensions[r].height = 42
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="CORE PROMPTS DISTRIBUTION BY DTC COMPONENT").font = SECTION_FONT
    r += 1
    header_row(ws, r, ["DTC Component", "Number of Prompts", "Required (Min)"])
    r += 1
    for comp, num, req in DTC:
        ws.cell(row=r, column=1, value=comp).font = NORMAL
        c = ws.cell(row=r, column=2, value=num)
        c.fill = YEL_FILL
        c.font = BOLD
        c.alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=3, value=req).alignment = Alignment(horizontal="center")
        for j in range(1, 4):
            ws.cell(row=r, column=j).border = BORDER
        r += 1


def build_detail(wb):
    ws = wb.create_sheet("Detailed AI Audit Log")
    ws.sheet_view.showGridLines = False
    headers = ["Entry #", "Prompt Type", "Stage/Component", "Problem/Context", "Prompt to AI",
               "AI Response (Summary)", "Human Delta & Reflection", "Evidence"]
    widths = [9, 17, 18, 32, 40, 40, 58, 30]
    style_title(ws, "DETAILED AI AUDIT LOG", len(headers))
    ws.cell(row=2, column=1,
            value="Only CORE prompts (Decision / Problem-Solving / Verification). "
                  "Each entry's Human Delta answers the 4 mandatory questions.").font = ITALIC
    header_row(ws, 3, headers, widths)
    r = 4
    for e in ENTRIES:
        for j, val in enumerate(e, start=1):
            c = ws.cell(row=r, column=j, value=val)
            c.font = NORMAL
            c.alignment = CENTER if j in (1, 2, 3) else TOPLEFT
            c.border = BORDER
        ws.row_dimensions[r].height = 170
        r += 1
    ws.freeze_panes = "A4"


def build_hallu(wb):
    ws = wb.create_sheet("Hallucination Detection Log")
    ws.sheet_view.showGridLines = False
    headers = ["Entry # (from Sheet 2)", "Hallucination Type", "AI's Claim", "Reality Check",
               "How Detected", "Corrective Action"]
    widths = [18, 26, 40, 40, 36, 40]
    style_title(ws, "HALLUCINATION DETECTION LOG (MANDATORY)", len(headers))
    ws.cell(row=2, column=1,
            value="DBI202 requires at least 3 detected hallucination cases.").font = Font(
        size=10, italic=True, bold=True, color="C00000", name="Calibri")
    header_row(ws, 3, headers, widths)
    r = 4
    for h in HALLU:
        label = "H" + str(r - 3) + " (Entry " + h[0] + ")"
        row_vals = [label] + h[1:]
        for j, val in enumerate(row_vals, start=1):
            c = ws.cell(row=r, column=j, value=val)
            c.font = NORMAL
            c.alignment = CENTER if j == 1 else TOPLEFT
            c.border = BORDER
        ws.row_dimensions[r].height = 110
        r += 1
    ws.freeze_panes = "A4"


def main():
    wb = Workbook()
    build_summary(wb)
    build_detail(wb)
    build_hallu(wb)
    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Lab4_AI_AuditLog.xlsx")
    wb.save(out)
    print("AI Audit Log saved to", out)


if __name__ == "__main__":
    main()
